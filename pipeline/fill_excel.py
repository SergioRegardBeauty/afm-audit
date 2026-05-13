"""
Lit les audits JSON par pays + flux et remplit les templates Excel BQ_1 (Cde Tel) et BQ_4 (SAV).
Mode dynamique : supporte N audits (insertion de lignes si N > 10).
Sort les fichiers remplis dans output/<COUNTRY>/.
"""
from __future__ import annotations

import json
import shutil
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(r"f:\WORK\Audit_Dialoga_AtlasForMen_2026")
AUDITS_DIR = ROOT / "audits_ia" / "v1"
OUTPUT_DIR = ROOT / "output"
TEMPLATE_CDE = ROOT / "BQ_1_CdesTel_10 (2).xlsx"
TEMPLATE_SAV = ROOT / "BQ_4_Sav_10_ (2).xlsx"

COUNTRIES = ["FR", "GB", "DE", "PL", "CZ"]

CDE_TEL_COLS = {
    "1_accueil":             (12, 13),
    "2_identification":      (14, 15),
    "3_climat":              (16, 17),
    "4_mise_en_attente":     (18, 19),
    "5_conclusion":          (20, 21),
    "6_prise_conges":        (22, 23),
    "7_personnalisation":    (24, 25),
    "8_pre_commande":        (26, 27),
    "9_commande":            (28, 29),
    "10_outils":             (30, 31),
    "11_coord_bancaires":    (32, 33),
    "12_rebond_commercial":  (34, 35),
    "13_objection":          (36, 37),
    "14_ecoute_active":      (38, 39),
    "15_accompagnement":     (40, 41),
}

SAV_COLS = {
    "1_accueil":                (12, 13),
    "2_identification":         (14, 15),
    "3_mise_en_attente":        (16, 17),
    "4_conclusion":             (18, 19),
    "5_prise_conges":           (20, 21),
    "6_personnalisation":       (22, 23),
    "7_climat":                 (24, 25),
    "8_maitrise_directivite":   (26, 27),
    "9_decouverte_reformulation": (28, 29),
    "10_pertinence_reponse":    (30, 31),
    "11_outils":                (32, 33),
    "12_codification":          (34, 35),
    "13_delai_traitement":      (36, 37),
    "14_experience_client":     (38, 39),
    "15_ecoute_empathie":       (40, 41),
    "16_enchantement":          (42, 43),
}

LANG_FROM_COUNTRY = {"FR": "fr", "GB": "en", "DE": "de", "PL": "pl", "CZ": "cs"}

FIRST_DATA_ROW = 24
LAST_TEMPLATE_DATA_ROW = 33   # template a 10 lignes : 24-33
SOUS_TOTAL_ROW = 34            # apres les 10 lignes data
HEADER_VALUE_COL = 4


def parse_note(val):
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.upper() in ("NA", "N/A", ""):
        return "NA"
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return s


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 65):
    """Recopie le style cellule par cellule de source_row vers target_row."""
    for c in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=c)
        tgt = ws.cell(row=target_row, column=c)
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.fill = copy(src.fill)
            tgt.border = copy(src.border)
            tgt.alignment = copy(src.alignment)
            tgt.number_format = src.number_format
            tgt.protection = copy(src.protection)


def fill_metadata_header(ws, country: str, lang: str, n_rows: int):
    today_str = date.today().strftime("%d/%m/%Y")
    ws.cell(row=4, column=HEADER_VALUE_COL, value=country)
    ws.cell(row=5, column=HEADER_VALUE_COL, value=lang)
    ws.cell(row=6, column=HEADER_VALUE_COL, value="Multi / IA")
    ws.cell(row=7, column=HEADER_VALUE_COL, value=today_str)
    ws.cell(row=8, column=HEADER_VALUE_COL, value=n_rows)
    ws.cell(row=9, column=HEADER_VALUE_COL, value="Claude IA v1 (audit transcript only)")
    ws.cell(row=11, column=HEADER_VALUE_COL, value=today_str)


def fill_audit_row(ws, row_idx: int, audit: dict, criteres_cols: dict, line_number: int):
    meta = audit.get("metadata", {})
    criteres = audit.get("criteres", {})
    totaux = audit.get("totaux", {})
    SI_list = audit.get("SI_detectees", [])

    # Col B (2) = numero de ligne
    ws.cell(row=row_idx, column=2, value=line_number)
    ws.cell(row=row_idx, column=3, value=meta.get("agent", ""))
    ws.cell(row=row_idx, column=4, value=meta.get("n_client", ""))
    ws.cell(row=row_idx, column=5, value=meta.get("pays", ""))
    ws.cell(row=row_idx, column=6, value=meta.get("n_cde_req", ""))
    ws.cell(row=row_idx, column=7, value=meta.get("n_tel", ""))
    ws.cell(row=row_idx, column=8, value=meta.get("DMT_DMC", ""))
    ws.cell(row=row_idx, column=9, value=meta.get("type_cde") or meta.get("type_demande", ""))
    ws.cell(row=row_idx, column=10, value=meta.get("demande_client", ""))
    ws.cell(row=row_idx, column=11, value=meta.get("n_appel", ""))

    for key, (note_col, bareme_col) in criteres_cols.items():
        c = criteres.get(key, {})
        note_val = parse_note(c.get("note"))
        bareme_val = parse_note(c.get("bareme"))
        notation = (c.get("notation") or "").upper()
        if "SI" in notation:
            note_val = 0
        elif notation in ("NA", "N/A"):
            note_val = "NA"
        ws.cell(row=row_idx, column=note_col, value=note_val)
        ws.cell(row=row_idx, column=bareme_col, value=bareme_val if bareme_val != "" else None)

    if totaux.get("note_zero_par_SI"):
        ws.cell(row=row_idx, column=52, value=0)
    else:
        ws.cell(row=row_idx, column=52, value=totaux.get("note_totale_sur_10"))

    comment = audit.get("commentaire_global", "")
    if SI_list:
        si_summary = " || SI: " + " ; ".join(
            f"{s.get('critere','?')} ({s.get('gravite','?')})" for s in SI_list
        )
        comment = comment + si_summary
    ws.cell(row=row_idx, column=53, value=comment)
    ws.cell(row=row_idx, column=58, value=audit.get("commentaire_traduit_fr", ""))


def load_audits_for(country: str, flow: str) -> list[dict]:
    folder = AUDITS_DIR / country / flow
    if not folder.exists():
        return []
    audits = []
    for p in sorted(folder.glob("*.json")):
        try:
            with p.open("r", encoding="utf-8") as f:
                audit = json.load(f)
            audit["_source_file"] = p.name
            audits.append(audit)
        except Exception as e:
            print(f"  WARN: skipped {p.name}: {e}")
    return audits


def fill_one_template(template_path: Path, output_path: Path, country: str,
                       audits: list[dict], criteres_cols: dict, sheet_name: str):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    lang = LANG_FROM_COUNTRY.get(country, "fr")
    N = len(audits)
    fill_metadata_header(ws, country, lang, N)

    if N == 0:
        wb.save(output_path)
        print(f"  + {output_path.name}: 0 audit (template vide)")
        return

    # Si plus de 10 audits, insere des lignes avant le sous total
    if N > 10:
        extra = N - 10
        # Insere extra lignes a partir de SOUS_TOTAL_ROW (qui devient SOUS_TOTAL_ROW + extra)
        ws.insert_rows(SOUS_TOTAL_ROW, amount=extra)
        # Copie le style d'une ligne data (24) vers les nouvelles lignes (34 a 33 + extra)
        for new_row in range(SOUS_TOTAL_ROW, SOUS_TOTAL_ROW + extra):
            copy_row_style(ws, source_row=FIRST_DATA_ROW, target_row=new_row)

    # Remplit les N lignes d'audit
    for i, audit in enumerate(audits):
        row_idx = FIRST_DATA_ROW + i
        fill_audit_row(ws, row_idx, audit, criteres_cols, line_number=i + 1)

    wb.save(output_path)
    print(f"  + {output_path.name}: {N} audits ecrits ({'10 + ' + str(N-10) + ' inseres' if N > 10 else str(N) + ' lignes template'})")


def write_country_readme(country: str, out_dir: Path, n_cde: int, n_sav: int):
    readme = out_dir / "README.md"
    today_str = date.today().strftime("%Y-%m-%d")
    content = f"""# Audit qualite IA — {country} — {today_str}

## Livrables
- `BQ_1_CdesTel_{country}.xlsx` — grille AFM Commande Telephone remplie ({n_cde} audits)
- `BQ_4_Sav_{country}.xlsx` — grille AFM SAV remplie ({n_sav} audits)
- `audits_json/` — JSON brut de chaque audit (1 fichier par appel)

## Methodologie
1. Echantillonnage exhaustif de transcripts substantiels (>3KB) parmi les
   exports CSV Dialoga, mois de janvier 2026.
2. Classification automatique Cde Tel / SAV par mots-cles multilingues.
3. Audit IA appel par appel selon les grilles BQ_1 / BQ_4 et le prompt
   `prompts/audit_cde_tel_v1.md` ou `prompts/audit_sav_v1.md`.
4. **Evaluation transcript-only** (pas de recording fourni) : les criteres lies
   au ton/sourire/empathie sont evalues sur les indices textuels uniquement
   (interjections, formulations positives, longueurs de pauses).

## Note pour la lecture du fichier Excel
Les templates ont ete etendus dynamiquement pour accueillir tous les audits du
mois (jusqu'a plusieurs centaines de lignes). Le bloc "sous total / total" en
bas a ete decale ; les formules originales (qui couvraient les 10 premieres
lignes) peuvent etre obsoletes. Recalculer manuellement ou utiliser un tableau
croise dynamique pour les analyses agregees.

## Limitations
- Les **Situations Inacceptables (SI)** detectees doivent etre **revues
  manuellement** avant toute action RH/coaching.
- L'identification du **prestataire** ne peut pas etre deduite du transcript :
  cellule "prestataire" laissee a "Multi / IA".
- Les criteres Climat / Empathie sont moins fiables sur transcript seul.
- Erreurs ASR frequentes sur "Atlas for Men" — flagges dans les justifications.
"""
    readme.write_text(content, encoding="utf-8")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=== Generation des Excel par pays (mode dynamique) ===\n")

    grand_total = {"cde_tel": 0, "sav": 0}

    for country in COUNTRIES:
        print(f"[{country}]")
        out_dir = OUTPUT_DIR / country
        out_dir.mkdir(parents=True, exist_ok=True)

        cde_audits = load_audits_for(country, "cde_tel")
        sav_audits = load_audits_for(country, "sav")
        grand_total["cde_tel"] += len(cde_audits)
        grand_total["sav"] += len(sav_audits)

        # Si le .xlsx est verrouille (ouvert dans Excel), on ecrit dans une version horodatee
        from datetime import datetime
        suffix = ""
        cde_out = out_dir / f"BQ_1_CdesTel_{country}.xlsx"
        sav_out = out_dir / f"BQ_4_Sav_{country}.xlsx"
        try:
            with cde_out.open("a"):
                pass
        except (PermissionError, OSError):
            ts = datetime.now().strftime("%H%M")
            suffix = f"_v{ts}"
            cde_out = out_dir / f"BQ_1_CdesTel_{country}{suffix}.xlsx"
            sav_out = out_dir / f"BQ_4_Sav_{country}{suffix}.xlsx"
            print(f"  ! fichier verrouille, ecriture dans {suffix}")

        fill_one_template(TEMPLATE_CDE, cde_out, country, cde_audits, CDE_TEL_COLS, "Cde tél")
        fill_one_template(TEMPLATE_SAV, sav_out, country, sav_audits, SAV_COLS, "SAV")

        # Copie JSONs to output for traceability
        json_dst = out_dir / "audits_json"
        json_dst.mkdir(parents=True, exist_ok=True)
        for flow, audits in (("cde_tel", cde_audits), ("sav", sav_audits)):
            flow_dst = json_dst / flow
            flow_dst.mkdir(parents=True, exist_ok=True)
            # Clean old jsons first
            for old in flow_dst.glob("*.json"):
                old.unlink()
            for audit in audits:
                src_name = audit.get("_source_file", "")
                src = AUDITS_DIR / country / flow / src_name
                if src.exists():
                    shutil.copy(src, flow_dst / src_name)

        write_country_readme(country, out_dir, len(cde_audits), len(sav_audits))
        print()

    print(f"=== TOTAL : {grand_total['cde_tel']} Cde Tel + {grand_total['sav']} SAV ===")


if __name__ == "__main__":
    main()
